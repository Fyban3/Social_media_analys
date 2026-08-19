import os
import glob
import tempfile
import io
from flask import Flask, render_template, request, jsonify, send_file, send_from_directory
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from db_storage import init_db, save_parsed_pdf, clear_all_data
from pdf_parser import parse_audit_pdf
from analytics import (
    get_overall_summary,
    get_personal_analytics,
    get_division_analytics,
    get_post_performance,
    get_post_detail,
    get_employee_detail
)
from pdf_generator import generate_pdf_report

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
app = Flask(__name__, static_folder="static", template_folder="templates")

# Initialize database schema on startup
init_db()

@app.route('/design/<path:filename>')
def serve_design(filename):
    """Serves design assets like official logos."""
    return send_from_directory(os.path.join(BASE_DIR, "design"), filename)

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/summary", methods=["GET"])
def api_summary():
    date_filter = request.args.get("date")
    summary = get_overall_summary(date_filter)
    return jsonify(summary)

@app.route("/api/personal", methods=["GET"])
def api_personal():
    date_filter = request.args.get("date")
    divisi_filter = request.args.get("divisi")
    search_query = request.args.get("search")
    data = get_personal_analytics(date_filter, divisi_filter, search_query)
    return jsonify(data)

@app.route("/api/divisions", methods=["GET"])
def api_divisions():
    date_filter = request.args.get("date")
    data = get_division_analytics(date_filter)
    return jsonify(data)

@app.route("/api/posts", methods=["GET"])
def api_posts():
    date_filter = request.args.get("date")
    data = get_post_performance(date_filter)
    return jsonify(data)

@app.route("/api/post/detail", methods=["GET"])
def api_post_detail():
    post_id = request.args.get("id")
    if not post_id:
        return jsonify({"error": "Post ID is required"}), 400
    data = get_post_detail(post_id)
    if not data:
        return jsonify({"error": "Post not found"}), 404
    return jsonify(data)

@app.route("/api/employee/detail", methods=["GET"])
def api_employee_detail():
    emp_name = request.args.get("name")
    date_filter = request.args.get("date")
    if not emp_name:
        return jsonify({"error": "Employee name required"}), 400
    data = get_employee_detail(emp_name, date_filter)
    return jsonify(data)

@app.route("/api/sync-downloads", methods=["POST", "GET"])
def api_sync_downloads():
    """
    Scans and ingests all audited PDFs from the user's Downloads folder.
    """
    downloads_dir = os.path.join(os.path.expanduser("~"), "Downloads")
    pdf_files = glob.glob(os.path.join(downloads_dir, "Rekap_Audit_Medsos_Pegawai_Kanwil_Kepri_*.pdf"))
    
    saved_count = 0
    errors = []
    
    for pdf_path in pdf_files:
        try:
            parsed = parse_audit_pdf(pdf_path)
            if parsed and parsed.get("employees"):
                save_parsed_pdf(parsed)
                saved_count += 1
        except Exception as e:
            errors.append(f"{os.path.basename(pdf_path)}: {str(e)}")
            
    return jsonify({
        "success": True,
        "message": f"Berhasil menyinkronkan {saved_count} file PDF audit dari folder Downloads.",
        "saved_count": saved_count,
        "total_scanned": len(pdf_files),
        "errors": errors
    })

@app.route("/api/upload", methods=["POST"])
def api_upload():
    if "files" not in request.files:
        return jsonify({"error": "No files uploaded"}), 400
        
    uploaded_files = request.files.getlist("files")
    saved_count = 0
    errors = []
    
    with tempfile.TemporaryDirectory() as temp_dir:
        for f in uploaded_files:
            if f.filename and f.filename.endswith(".pdf"):
                temp_path = os.path.join(temp_dir, f.filename)
                f.save(temp_path)
                try:
                    parsed = parse_audit_pdf(temp_path)
                    save_parsed_pdf(parsed)
                    saved_count += 1
                except Exception as e:
                    errors.append(f"{f.filename}: {str(e)}")
                    
    return jsonify({
        "success": True,
        "saved_count": saved_count,
        "errors": errors
    })

@app.route("/api/import-folder", methods=["POST"])
def api_import_folder():
    req_data = request.get_json() or {}
    folder_path = req_data.get("folder_path", os.path.join(os.path.expanduser("~"), "Downloads"))
    
    if not os.path.exists(folder_path):
        return jsonify({"error": f"Directory path '{folder_path}' does not exist."}), 400
        
    pdf_files = glob.glob(os.path.join(folder_path, "*.pdf"))
    if not pdf_files:
        return jsonify({"error": f"No PDF files found in '{folder_path}'."}), 400
        
    saved_count = 0
    errors = []
    for pdf_path in pdf_files:
        try:
            parsed = parse_audit_pdf(pdf_path)
            if parsed and parsed.get("employees"):
                save_parsed_pdf(parsed)
                saved_count += 1
        except Exception as e:
            errors.append(f"{os.path.basename(pdf_path)}: {str(e)}")
            
    return jsonify({
        "success": True,
        "folder_path": folder_path,
        "saved_count": saved_count,
        "errors": errors
    })

@app.route("/api/post/export-excel", methods=["GET"])
def api_export_single_post_excel():
    post_id = request.args.get("id")
    if not post_id:
        return jsonify({"error": "Post ID required"}), 400
        
    data = get_post_detail(post_id)
    if not data:
        return jsonify({"error": "Post not found"}), 404
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekap Audit Postingan"
    ws.views.sheetView[0].showGridLines = True
    
    # Styling
    font_title = Font(name="Calibri", size=13, bold=True, color="0F172A")
    font_meta = Font(name="Calibri", size=9.5, bold=True, color="334155")
    font_header = Font(name="Calibri", size=9.5, bold=True, color="FFFFFF")
    font_div = Font(name="Calibri", size=10, bold=True, color="000000")
    font_data = Font(name="Calibri", size=9.5, color="000000")
    
    fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    fill_tu = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_yankum = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    fill_pp = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
    fill_kakanwil = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center')
    
    # Title Block
    ws['A1'] = "REKAPITULASI AUDIT INTERAKSI MEDIA SOSIAL PEGAWAI"
    ws['A1'].font = font_title
    ws.merge_cells('A1:G1')
    
    ws['A2'] = "KANTOR WILAYAH KEMENTERIAN HUKUM KEPULAUAN RIAU"
    ws['A2'].font = font_meta
    ws.merge_cells('A2:G2')
    
    ws['A4'] = f"JUDUL POST IG : {data['ig_title']} (URL: {data['ig_url']})"
    ws['A4'].font = font_meta
    ws.merge_cells('A4:G4')
    
    ws['A5'] = f"JUDUL POST FB : {data['fb_title']} (URL: {data['fb_url']})"
    ws['A5'].font = font_meta
    ws.merge_cells('A5:G5')
    
    ws['A6'] = f"TANGGAL AUDIT : {data['audit_date']}   |   WAKTU EKSPOR: {data['export_time']}"
    ws['A6'].font = font_meta
    ws.merge_cells('A6:G6')
    
    # Table Header (Row 8-9)
    ws.merge_cells('A8:A9')
    ws['A8'] = "NO"
    ws.merge_cells('B8:B9')
    ws['B8'] = "NAMA PEGAWAI"
    ws.merge_cells('C8:C9')
    ws['C8'] = "JABATAN"
    
    ws.merge_cells('D8:E8')
    ws['D8'] = "INSTAGRAM"
    ws.merge_cells('F8:G8')
    ws['F8'] = "FACEBOOK"
    
    ws['D9'] = "LIKE"
    ws['E9'] = "KOMEN"
    ws['F9'] = "LIKE"
    ws['G9'] = "KOMEN"
    
    for r in range(8, 10):
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = thin_border
            
    current_row = 10
    for div, items in data['grouped_by_divisi'].items():
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        div_cell = ws.cell(row=current_row, column=1, value=f"DIVISI: {div}")
        div_cell.font = font_div
        div_cell.alignment = align_left
        
        fill_c = fill_tu
        if "KEPALA KANTOR WILAYAH" in div:
            fill_c = fill_kakanwil
        elif "PELAYANAN HUKUM" in div:
            fill_c = fill_yankum
        elif "PERATURAN PERUNDANG" in div:
            fill_c = fill_pp
            
        for c in range(1, 8):
            ws.cell(row=current_row, column=c).fill = fill_c
            ws.cell(row=current_row, column=c).border = thin_border
            
        current_row += 1
        
        for idx, itm in enumerate(items, 1):
            ws.cell(row=current_row, column=1, value=idx).alignment = align_center
            ws.cell(row=current_row, column=2, value=itm['nama']).alignment = align_left
            ws.cell(row=current_row, column=3, value=itm['jabatan']).alignment = align_left
            
            def sym(v):
                return "✅" if v == "SUDAH" else ("❌" if v == "BELUM" else "-")
                
            ws.cell(row=current_row, column=4, value=sym(itm['ig_like'])).alignment = align_center
            ws.cell(row=current_row, column=5, value=sym(itm['ig_komen'])).alignment = align_center
            ws.cell(row=current_row, column=6, value=sym(itm['fb_like'])).alignment = align_center
            ws.cell(row=current_row, column=7, value=sym(itm['fb_komen'])).alignment = align_center
            
            for c in range(1, 8):
                ws.cell(row=current_row, column=c).font = font_data
                ws.cell(row=current_row, column=c).border = thin_border
                
            current_row += 1
            
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 44
    for c in ['D', 'E', 'F', 'G']:
        ws.column_dimensions[c].width = 14
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Rekap_Audit_Post_{data['audit_date']}_{post_id}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )

@app.route("/api/export-excel", methods=["GET"])
def api_export_excel():
    date_filter = request.args.get("date")
    divisi_filter = request.args.get("divisi")
    search_query = request.args.get("search")
    
    personal_data = get_personal_analytics(date_filter, divisi_filter, search_query)
    
    df_data = []
    for idx, emp in enumerate(personal_data, 1):
        df_data.append({
            "No": idx,
            "Nama Pegawai": emp['nama'],
            "Jabatan": emp['jabatan'],
            "Divisi": emp['divisi'],
            "Jumlah Post Audited": emp['total_posts'],
            "IG Like": emp['ig_like'],
            "IG Komen": emp['ig_komen'],
            "FB Like": emp['fb_like'],
            "FB Komen": emp['fb_komen'],
            "Total Like (IG+FB)": emp['total_like'],
            "Total Komen (IG+FB)": emp['total_komen'],
            "Total Interaksi": emp['total_interaction'],
            "Persentase Kepatuhan Like (%)": emp['like_compliance']
        })
        
    df = pd.DataFrame(df_data)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Rekap Analisis Personal')
    output.seek(0)
    
    export_filename = f"Rekap_Analisis_Medsos_Personal_{date_filter or 'All'}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=export_filename
    )

@app.route("/api/export-pdf", methods=["GET"])
def api_export_pdf():
    date_filter = request.args.get("date")
    divisi_filter = request.args.get("divisi")
    search_query = request.args.get("search")
    
    pdf_buffer = generate_pdf_report(date_filter, divisi_filter, search_query)
    export_filename = f"Laporan_Analisis_Medsos_{date_filter or 'All'}.pdf"
    
    return send_file(
        pdf_buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=export_filename
    )

@app.route("/api/clear", methods=["POST"])
def api_clear():
    clear_all_data()
    return jsonify({"success": True, "message": "All data cleared successfully."})

def auto_ingest_downloads():
    """Auto-scans and ingests all audited PDFs from user's Downloads on startup."""
    downloads_dir = os.path.join(os.path.expanduser("~"), "Downloads")
    pdf_files = glob.glob(os.path.join(downloads_dir, "Rekap_Audit_Medsos_Pegawai_Kanwil_Kepri_*.pdf"))
    if pdf_files:
        print(f"[*] Found {len(pdf_files)} audit PDFs in Downloads folder. Synchronizing database...")
        ingested = 0
        for p in pdf_files:
            try:
                parsed = parse_audit_pdf(p)
                if parsed and parsed.get("employees"):
                    save_parsed_pdf(parsed)
                    ingested += 1
            except Exception as e:
                print(f"[!] Error parsing {p}: {e}")
        print(f"[OK] Ingested {ingested} PDF audits from Downloads folder into database!")

if __name__ == "__main__":
    auto_ingest_downloads()
    print("=" * 60)
    print("[OK] Social Media Audit & Analytics System running on http://127.0.0.1:5050")
    print("=" * 60)
    app.run(host="0.0.0.0", port=5050, debug=False)
